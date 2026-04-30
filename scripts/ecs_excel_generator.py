#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 生成器 - 生成阿里云资源清单 Excel 文件
严格按照 REQUIREMENTS.md 的格式要求实现

注意：场景一（CSV 报价）传入 include_instance_id=True，会在产品名称和产品描述之间插入"实例 ID"列；
      场景二（文本报价）传入 include_instance_id=False（默认），格式不变。
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import MergeCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


class ExcelGenerator:
    """Excel 生成器类"""
    
    # 颜色定义
    SKY_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # 天蓝色
    RED_FONT = Font(color="FF0000")  # 红色字体
    
    # 字体样式
    DEFAULT_FONT = Font(name="微软雅黑", size=11, bold=True)
    
    # 对齐方式
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    WRAP_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 自动换行
    
    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 输出目录（优先读取环境变量，由入口脚本从 config.json 设置）
    OUTPUT_DIR = os.environ.get("ALIYUN_SKILL_OUTPUT_DIR", "/root/.openclaw/workspace/download/")
    
    def __init__(self, customer_name: str = "", include_instance_id: bool = False):
        """
        初始化 Excel 生成器
        
        Args:
            customer_name: 客户名称
            include_instance_id: 是否在 Excel 中包含"实例 ID"列（仅场景一 CSV 报价时设为 True）
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "资源清单"
        self.customer_name = customer_name
        self.include_instance_id = include_instance_id
        self.data_rows = []
        
    def add_data_row(self, product_name: str, product_desc: str, region: str, 
                     quantity: int = 1, price_1y_list: float = None, 
                     price_1y_discount: float = None, price_3y_list: float = None,
                     price_3y_discount: float = None, remark: str = "", 
                     is_error: bool = False, instance_id: str = ""):
        """
        添加数据行
        
        Args:
            product_name: 产品名称
            instance_id: 实例 ID（仅场景一 CSV 报价时生效）
            product_desc: 产品描述
            region: 地域
            quantity: 数量
            price_1y_list: 官网目录价（1 年）
            price_1y_discount: 官网折扣价（1 年）
            price_3y_list: 官网目录价（3 年）
            price_3y_discount: 官网折扣价（3 年）
            remark: 备注
            is_error: 是否为错误行（询价失败）
        """
        self.data_rows.append({
            "product_name": product_name,
            "instance_id": instance_id,
            "product_desc": product_desc,
            "region": region,
            "quantity": quantity,
            "price_1y_list": price_1y_list,
            "price_1y_discount": price_1y_discount,
            "price_3y_list": price_3y_list,
            "price_3y_discount": price_3y_discount,
            "remark": remark,
            "is_error": is_error
        })
    
    def _generate_filename(self) -> str:
        """生成文件名：阿里云资源清单 YYMMDD-HHMMSS.xlsx"""
        timestamp = datetime.now().strftime("%y%m%d-%H%M%S")
        return f"阿里云资源清单{timestamp}.xlsx"
    
    def _apply_header_style(self, row: int, col_start: int, col_end: int):
        """应用标头行样式"""
        for col in range(col_start, col_end + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.font = self.DEFAULT_FONT
            cell.fill = self.SKY_BLUE_FILL
            cell.alignment = self.CENTER_ALIGNMENT
            cell.border = self.THIN_BORDER
    
    def _apply_data_style(self, row: int, col_start: int, col_end: int, is_error: bool = False):
        """应用数据行样式"""
        for col in range(col_start, col_end + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.font = self.DEFAULT_FONT
            cell.alignment = self.CENTER_ALIGNMENT
            cell.border = self.THIN_BORDER
            # A 列背景色
            if col == col_start:
                cell.fill = self.SKY_BLUE_FILL
            # 错误行备注列红色字体
            if is_error and col == col_end:
                cell.font = Font(name="微软雅黑", size=11, bold=True, color="FF0000")
    
    def _format_currency(self, value: float) -> str:
        """格式化货币：￥XX.XX"""
        if value is None:
            return ""
        return f"￥{value:,.2f}"
    
    def generate(self) -> str:
        """
        生成 Excel 文件
        
        Returns:
            生成的文件路径
        """
        ws = self.ws
        
        # 确保输出目录存在
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)
        
        quote_date = datetime.now().strftime("%Y年%m月%d日")
        
        # 根据 include_instance_id 决定列数和布局
        if self.include_instance_id:
            # 场景一（CSV 报价）：10列 = 产品名称 | 实例ID | 产品描述 | 地域 | 数量 | 价格x4 | 备注
            num_cols = 10
            last_col_letter = 'J'
            headers = [
                "产品名称", "实例 ID", "产品描述", "地域", "数量",
                "官网目录价（元/1 年）", "官网折扣价（元/1 年）",
                "官网目录价（元/3 年）", "官网折扣价（元/3 年）",
                "备注"
            ]
            column_widths = [15, 28, 50, 18, 10, 25, 25, 25, 25, 35]
            # 数据行列位置
            col_product_name = 1
            col_instance_id = 2
            col_product_desc = 3
            col_region = 4
            col_quantity = 5
            col_price_1y_list = 6
            col_price_1y_discount = 7
            col_price_3y_list = 8
            col_price_3y_discount = 9
            col_remark = 10
            # 合计行价格列范围
            price_start_col = 6
            price_end_col = 9
        else:
            # 场景二（文本报价）：9列 = 产品名称 | 产品描述 | 地域 | 数量 | 价格x4 | 备注
            num_cols = 9
            last_col_letter = 'I'
            headers = [
                "产品名称", "产品描述", "地域", "数量",
                "官网目录价（元/1 年）", "官网折扣价（元/1 年）",
                "官网目录价（元/3 年）", "官网折扣价（元/3 年）",
                "备注"
            ]
            column_widths = [15, 50, 18, 10, 25, 25, 25, 25, 35]
            # 数据行列位置
            col_product_name = 1
            col_instance_id = None  # 不存在
            col_product_desc = 2
            col_region = 3
            col_quantity = 4
            col_price_1y_list = 5
            col_price_1y_discount = 6
            col_price_3y_list = 7
            col_price_3y_discount = 8
            col_remark = 9
            # 合计行价格列范围
            price_start_col = 5
            price_end_col = 8
        
        quote_date = datetime.now().strftime("%Y年%m月%d日")
        
        # === 第一行：标题行 ===
        date_col = num_cols
        # 合并单元格范围（A 到最后一列前一个）
        if num_cols > 1:
            merge_end_col = num_cols - 1
        else:
            merge_end_col = 1
        merge_end_letter = get_column_letter(merge_end_col)
        ws.merge_cells(f'A1:{merge_end_letter}1')
        cell_a1 = ws.cell(row=1, column=1)
        cell_a1.value = f"客户名称：{self.customer_name}" if self.customer_name else "客户名称："
        cell_a1.font = self.DEFAULT_FONT
        cell_a1.alignment = self.LEFT_ALIGNMENT
        
        ws.cell(row=1, column=date_col).value = f"报价日期：{quote_date}"
        ws.cell(row=1, column=date_col).font = self.DEFAULT_FONT
        ws.cell(row=1, column=date_col).alignment = self.CENTER_ALIGNMENT
        
        for col in range(1, num_cols + 1):
            ws.cell(row=1, column=col).border = self.THIN_BORDER
        
        # === 第二行：标头行 ===
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            self._apply_header_style(2, col, col)
        
        # === 第三行起：数据行 ===
        current_row = 3
        for data in self.data_rows:
            ws.cell(row=current_row, column=col_product_name, value=data["product_name"])
            if self.include_instance_id:
                ws.cell(row=current_row, column=col_instance_id, value=data["instance_id"])
            ws.cell(row=current_row, column=col_product_desc, value=data["product_desc"])
            ws.cell(row=current_row, column=col_region, value=data["region"])
            ws.cell(row=current_row, column=col_quantity, value=data["quantity"])
            ws.cell(row=current_row, column=col_price_1y_list, value=data["price_1y_list"])
            ws.cell(row=current_row, column=col_price_1y_discount, value=data["price_1y_discount"])
            ws.cell(row=current_row, column=col_price_3y_list, value=data["price_3y_list"])
            ws.cell(row=current_row, column=col_price_3y_discount, value=data["price_3y_discount"])
            ws.cell(row=current_row, column=col_remark, value=data["remark"])
            
            self._apply_data_style(current_row, 1, num_cols, is_error=data["is_error"])
            
            ws.cell(row=current_row, column=col_product_desc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=current_row, column=col_remark).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            ws.row_dimensions[current_row].height = 100
            
            current_row += 1
        
        # === 合计行 ===
        total_row = current_row
        ws.cell(row=total_row, column=1, value="合计")
        self._apply_data_style(total_row, 1, num_cols)
        
        for col in range(price_start_col, price_end_col + 1):
            is_1y = (col <= price_start_col + 1)
            is_list = (col % 2 == (0 if self.include_instance_id else 1))
            key = f"price_{'1' if is_1y else '3'}y_{'list' if is_list else 'discount'}"
            total = sum(row[key] for row in self.data_rows if row[key] is not None)
            ws.cell(row=total_row, column=col, value=total)
        
        # === 备注行 ===
        note_row = total_row + 1
        ws.merge_cells(f'A{note_row}:{last_col_letter}{note_row}')
        cell_note = ws.cell(row=note_row, column=1)
        
        cell_note.value = CellRichText(
            TextBlock(InlineFont(rFont='微软雅黑', sz=11, b=True), "备注："),
            TextBlock(InlineFont(rFont='微软雅黑', sz=11, b=True, color='FFFF0000'), "产品价格可能有波动，以官网实际价格为准")
        )
        cell_note.alignment = self.CENTER_ALIGNMENT
        
        for col in range(1, num_cols + 1):
            ws.cell(row=note_row, column=col).border = self.THIN_BORDER
        
        ws.row_dimensions[note_row].height = 30
        
        # === 设置列宽 ===
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # === 保存文件 ===
        filename = self._generate_filename()
        filepath = os.path.join(self.OUTPUT_DIR, filename)
        self.wb.save(filepath)
        
        return filepath
    
    def cleanup(self, filepath: str):
        """清理本地文件"""
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
                print(f"已删除本地文件：{filepath}")
        except Exception as e:
            print(f"删除文件失败：{e}")


def create_product_desc(instance_spec: str, image: str, system_disk: str, 
                        data_disk: str, bandwidth: str) -> str:
    """创建产品描述文本"""
    desc_parts = []
    if instance_spec:
        desc_parts.append(f"实例规格：{instance_spec}")
    if image:
        desc_parts.append(f"镜像：{image}")
    if system_disk:
        desc_parts.append(f"系统盘：{system_disk}")
    if data_disk:
        desc_parts.append(f"数据盘：{data_disk}")
    if bandwidth:
        desc_parts.append(f"公网带宽：{bandwidth}")
    return "\n".join(desc_parts)


if __name__ == "__main__":
    # 测试场景二（无实例ID列）
    print("测试场景二（无实例ID）...")
    generator = ExcelGenerator(customer_name="测试客户", include_instance_id=False)
    generator.add_data_row(
        product_name="ECS",
        product_desc="实例规格：ecs.c9i.xlarge (4 vCPU 8 GiB)\n镜像：公共免费镜像\n系统盘：ESSD 云盘 PL0 40GiB",
        region="杭州（cn-hangzhou）",
        quantity=1,
        price_1y_list=1234.56,
        price_1y_discount=987.65,
        price_3y_list=3703.68,
        price_3y_discount=2962.95,
        remark="首年优惠"
    )
    filepath = generator.generate()
    print(f"场景二 Excel：{filepath}")
    
    # 测试场景一（有实例ID列）
    print("\n测试场景一（有实例ID）...")
    generator2 = ExcelGenerator(customer_name="测试客户", include_instance_id=True)
    generator2.add_data_row(
        product_name="ECS",
        instance_id="i-uf6abc123",
        product_desc="实例规格：ecs.g9i.2xlarge (8 vCPU 32 GiB)\n镜像：公共免费镜像\n系统盘：ESSD 云盘 PL0 50GiB",
        region="上海（cn-shanghai）",
        quantity=1,
        price_1y_list=2345.67,
        price_1y_discount=1876.54,
        price_3y_list=7037.01,
        price_3y_discount=5629.62,
        remark="新客优惠"
    )
    filepath2 = generator2.generate()
    print(f"场景一 Excel：{filepath2}")
