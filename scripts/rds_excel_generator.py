#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RDS Excel 生成器 - 生成阿里云 RDS 资源清单 Excel 文件
复用 ECS 的 Excel 生成器,仅调整产品名称为 RDS
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import MergeCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


class RDSExcelGenerator:
    """RDS Excel 生成器类"""

    # 颜色定义
    SKY_BLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # 天蓝色
    RED_FONT = Font(color="FF0000")  # 红色字体

    # 字体样式
    DEFAULT_FONT = Font(name="微软雅黑", size=11, bold=True)

    # 对齐方式
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
    WRAP_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 自动换行

    # 边框样式
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 输出目录
    OUTPUT_DIR = "/root/.openclaw/workspace/download/"

    def __init__(self, customer_name: str = "", include_instance_id: bool = True):
        """
        初始化 RDS Excel 生成器
        
        Args:
            customer_name: 客户名称
            include_instance_id: 是否包含"实例 ID"列（默认 True）
        """
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "资源清单"
        self.customer_name = customer_name
        self.include_instance_id = include_instance_id
        self.data_rows = []

    def add_data_row(self, product_name: str, product_desc: str, region: str,
                     quantity: int = 1, price_1y_list: float = None,
                     price_1y_discount: float = None, price_3y_list: float = None,
                     price_3y_discount: float = None, remark: str = "",
                     is_error: bool = False, instance_id: str = ""):
        """
        添加数据行

        Args:
            product_name: 产品名称
            instance_id: 实例 ID
            product_desc: 产品描述
            region: 地域
            quantity: 数量
            price_1y_list: 官网目录价（1 年）
            price_1y_discount: 官网折扣价（1 年）
            price_3y_list: 官网目录价（3 年）
            price_3y_discount: 官网折扣价（3 年）
            remark: 备注
            is_error: 是否为错误行（询价失败）
        """
        self.data_rows.append({
            "product_name": product_name,
            "instance_id": instance_id,
            "product_desc": product_desc,
            "region": region,
            "quantity": quantity,
            "price_1y_list": price_1y_list,
            "price_1y_discount": price_1y_discount,
            "price_3y_list": price_3y_list,
            "price_3y_discount": price_3y_discount,
            "remark": remark,
            "is_error": is_error
        })

    def _generate_filename(self) -> str:
        """生成文件名:阿里云资源清单 YYMMDD-HHMMSS.xlsx"""
        timestamp = datetime.now().strftime("%y%m%d-%H%M%S")
        return f"阿里云资源清单{timestamp}.xlsx"

    def _apply_header_style(self, row: int, col_start: int, col_end: int):
        """应用标头行样式"""
        for col in range(col_start, col_end + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.font = self.DEFAULT_FONT
            cell.fill = self.SKY_BLUE_FILL
            cell.alignment = self.CENTER_ALIGNMENT
            cell.border = self.THIN_BORDER  # 添加边框

    def _apply_data_style(self, row: int, col_start: int, col_end: int, is_error: bool = False):
        """应用数据行样式"""
        for col in range(col_start, col_end + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.font = self.DEFAULT_FONT
            cell.alignment = self.CENTER_ALIGNMENT
            cell.border = self.THIN_BORDER  # 添加边框
            # A 列背景色(从第二行到合计行)
            if col == col_start:
                cell.fill = self.SKY_BLUE_FILL
            # 错误行备注列红色字体
            if is_error and col == col_end:
                cell.font = Font(name="微软雅黑", size=11, bold=True, color="FF0000")

    def _format_currency(self, value: float) -> str:
        """格式化货币:¥XX.XX"""
        if value is None:
            return ""
        return f"¥{value:,.2f}"

    def generate(self) -> str:
        """
        生成 Excel 文件

        Returns:
            生成的文件路径
        """
        ws = self.ws

        # 确保输出目录存在
        os.makedirs(self.OUTPUT_DIR, exist_ok=True)

        # 根据 include_instance_id 决定列数和布局
        if self.include_instance_id:
            num_cols = 10
            last_col_letter = 'J'
            headers = [
                "产品名称", "实例 ID", "产品描述", "地域", "数量",
                "官网目录价（元/1 年）", "官网折扣价（元/1 年）",
                "官网目录价（元/3 年）", "官网折扣价（元/3 年）",
                "备注"
            ]
            col_product_name = 1
            col_instance_id = 2
            col_product_desc = 3
            col_region = 4
            col_quantity = 5
            col_price_1y_list = 6
            col_price_1y_discount = 7
            col_price_3y_list = 8
            col_price_3y_discount = 9
            col_remark = 10
            date_col = num_cols
            merge_end_col = num_cols - 1
            column_widths = [15, 28, 50, 18, 10, 25, 25, 25, 25, 35]
            price_start_col = 6
            price_end_col = 9
        else:
            num_cols = 9
            last_col_letter = 'I'
            headers = [
                "产品名称", "产品描述", "地域", "数量",
                "官网目录价（元/1 年）", "官网折扣价（元/1 年）",
                "官网目录价（元/3 年）", "官网折扣价（元/3 年）",
                "备注"
            ]
            col_product_name = 1
            col_instance_id = None
            col_product_desc = 2
            col_region = 3
            col_quantity = 4
            col_price_1y_list = 5
            col_price_1y_discount = 6
            col_price_3y_list = 7
            col_price_3y_discount = 8
            col_remark = 9
            date_col = num_cols
            merge_end_col = num_cols - 1
            column_widths = [15, 50, 18, 10, 25, 25, 25, 25, 35]
            price_start_col = 5
            price_end_col = 8
        
        quote_date = datetime.now().strftime("%Y年%m月%d日")
        
        # === 第一行：标题行 ===
        merge_end_letter = get_column_letter(merge_end_col)
        ws.merge_cells(f'A1:{merge_end_letter}1')
        cell_a1 = ws.cell(row=1, column=1)
        cell_a1.value = f"客户名称：{self.customer_name}" if self.customer_name else "客户名称："
        cell_a1.font = self.DEFAULT_FONT
        cell_a1.alignment = self.LEFT_ALIGNMENT
        
        ws.cell(row=1, column=date_col).value = f"报价日期：{quote_date}"
        ws.cell(row=1, column=date_col).font = self.DEFAULT_FONT
        ws.cell(row=1, column=date_col).alignment = self.CENTER_ALIGNMENT
        
        for col in range(1, num_cols + 1):
            ws.cell(row=1, column=col).border = self.THIN_BORDER

        # === 第二行：标头行 ===
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            self._apply_header_style(2, col, col)

        # === 第三行起：数据行 ===
        current_row = 3
        for data in self.data_rows:
            ws.cell(row=current_row, column=col_product_name, value=data["product_name"])
            if self.include_instance_id:
                ws.cell(row=current_row, column=col_instance_id, value=data["instance_id"])
            ws.cell(row=current_row, column=col_product_desc, value=data["product_desc"])
            ws.cell(row=current_row, column=col_region, value=data["region"])
            ws.cell(row=current_row, column=col_quantity, value=data["quantity"])
            ws.cell(row=current_row, column=col_price_1y_list, value=data["price_1y_list"])
            ws.cell(row=current_row, column=col_price_1y_discount, value=data["price_1y_discount"])
            ws.cell(row=current_row, column=col_price_3y_list, value=data["price_3y_list"])
            ws.cell(row=current_row, column=col_price_3y_discount, value=data["price_3y_discount"])
            ws.cell(row=current_row, column=col_remark, value=data["remark"])
            
            self._apply_data_style(current_row, 1, num_cols, is_error=data["is_error"])
            
            ws.cell(row=current_row, column=col_product_desc).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.cell(row=current_row, column=col_remark).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            ws.row_dimensions[current_row].height = 100
            
            current_row += 1

        # === 合计行（倒数第二行）===
        total_row = current_row
        ws.cell(row=total_row, column=1, value="合计")
        self._apply_data_style(total_row, 1, num_cols)
        
        # 计算合计值
        for col in range(price_start_col, price_end_col + 1):
            is_1y = (col <= price_start_col + 1)
            is_list = (col % 2 == 0)
            key = f"price_{'1' if is_1y else '3'}y_{'list' if is_list else 'discount'}"
            total = sum(row[key] for row in self.data_rows if row[key] is not None)
            ws.cell(row=total_row, column=col, value=total)

        # === 最后一行：备注行 ===
        note_row = total_row + 1
        ws.merge_cells(f'A{note_row}:{last_col_letter}{note_row}')
        cell_note = ws.cell(row=note_row, column=1)

        # 使用富文本实现部分文字红色
        # "备注:"黑色 + "产品价格可能有波动,以官网实际价格为准"红色
        cell_note.value = CellRichText(
            TextBlock(InlineFont(rFont='微软雅黑', sz=11, b=True), "备注:"),
            TextBlock(InlineFont(rFont='微软雅黑', sz=11, b=True, color='FFFF0000'), "产品价格可能有波动,以官网实际价格为准")
        )
        # 居中对齐
        cell_note.alignment = self.CENTER_ALIGNMENT

        # 备注行所有单元格添加边框
        for col in range(1, num_cols + 1):
            ws.cell(row=note_row, column=col).border = self.THIN_BORDER

        ws.row_dimensions[note_row].height = 30

        # === 设置列宽 ===
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # === 保存文件 ===
        filename = self._generate_filename()
        filepath = os.path.join(self.OUTPUT_DIR, filename)
        self.wb.save(filepath)

        return filepath

    def cleanup(self, filepath: str):
        """
        清理本地文件

        Args:
            filepath: 文件路径
        """
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
                print(f"已删除本地文件:{filepath}")
        except Exception as e:
            print(f"删除文件失败:{e}")


def create_rds_product_desc(engine: str, engine_version: str, series: str,
                            storage_type: str, spec_desc: str, instance_class: str,
                            storage: int) -> str:
    """
    创建 RDS 产品描述文本

    格式要求:
    引擎:mysql 8.0
    产品系列:高可用系列
    存储类型:高性能本地盘
    实例规格:4 核 8GB(通用型)rds.mysql.s3.large
    存储空间:500

    Args:
        engine: 引擎名称(应小写)
        engine_version: 引擎版本
        series: 产品系列
        storage_type: 存储类型
        spec_desc: 规格描述(如 "4 核 8GB(通用型)")
        instance_class: 实例规格代码
        storage: 存储空间(GB)

    Returns:
        格式化后的产品描述
    """
    desc_parts = []

    # 引擎:<引擎><版本号>(引擎名称小写)
    desc_parts.append(f"引擎:{engine.lower()} {engine_version}")

    # 产品系列
    if series:
        desc_parts.append(f"产品系列:{series}")

    # 存储类型
    if storage_type:
        desc_parts.append(f"存储类型:{storage_type}")

    # 实例规格:<规格描述><规格>
    if spec_desc and instance_class:
        desc_parts.append(f"实例规格:{spec_desc} {instance_class}")
    elif instance_class:
        desc_parts.append(f"实例规格:{instance_class}")

    # 存储空间:<空间>GB
    desc_parts.append(f"存储空间:{storage}GB")

    return "\n".join(desc_parts)
