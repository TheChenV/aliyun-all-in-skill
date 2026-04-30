#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OSS Excel 生成器 - 生成 OSS 使用统计 Excel 文件
"""

import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class OSSExcelGenerator:
    """OSS Excel 生成器"""
    
    def __init__(self):
        """初始化 Excel 生成器"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "OSS 使用统计"
        
        # 设置列宽
        self.ws.column_dimensions['A'].width = 15  # 地域
        self.ws.column_dimensions['B'].width = 15  # 存储类型
        self.ws.column_dimensions['C'].width = 12  # 冗余类型
        self.ws.column_dimensions['D'].width = 18  # 实际存储
        self.ws.column_dimensions['E'].width = 18  # 计费存储
        
        # 数据行
        self._rows: List[dict] = []
        
        # 样式定义
        self._setup_styles()
    
    def _setup_styles(self):
        """设置样式"""
        # 标头样式：淡蓝色背景 (#D9E1F2)，微软雅黑，加粗，居中
        self.header_font = Font(name='Microsoft YaHei', bold=True, size=11)
        self.header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 数据行样式：微软雅黑，加粗，居中，所有线框
        self.data_font = Font(name='Microsoft YaHei', bold=True, size=11)
        self.data_alignment = Alignment(horizontal='center', vertical='center')
        
        # 边框
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def add_row(self, region: str, storage_class: str,
                redundancy_type: str, real_storage: float, billing_storage: float):
        """
        添加数据行
        
        Args:
            region: 地域
            storage_class: 存储类型
            redundancy_type: 冗余类型
            real_storage: 实际存储量（字节）
            billing_storage: 计费存储量（字节）
        """
        self._rows.append({
            'region': region,
            'storage_class': storage_class,
            'redundancy_type': redundancy_type,
            'real_storage': real_storage,
            'billing_storage': billing_storage
        })
    
    def _format_storage(self, bytes_value: float) -> str:
        """
        格式化存储大小（XX.XX GB）
        
        Args:
            bytes_value: 字节数
            
        Returns:
            格式化后的字符串（XX.XX GB）
        """
        gb_value = bytes_value / (1024 ** 3)
        return f"{gb_value:.2f} GB"
    
    def _translate_storage_class(self, storage_class: str) -> str:
        """
        翻译存储类型为中文
        
        Args:
            storage_class: 存储类型英文
            
        Returns:
            存储类型中文
        """
        mapping = {
            'Standard': '标准存储',
            'IA': '低频访问',
            'Archive': '归档存储',
            'ColdArchive': '冷归档存储'
        }
        return mapping.get(storage_class, storage_class)
    
    def _translate_redundancy_type(self, redundancy_type: str) -> str:
        """
        翻译冗余类型为中文
        
        Args:
            redundancy_type: 冗余类型英文
            
        Returns:
            冗余类型中文
        """
        mapping = {
            'LRS': '本地冗余',
            'ZRS': '同城冗余'
        }
        return mapping.get(redundancy_type, redundancy_type)
    
    def generate(self) -> str:
        """
        生成 Excel 文件
        
        Returns:
            文件路径
        """
        # 写入标头
        headers = ['地域', '存储类型', '冗余类型', '实际存储大小', '计费存储大小']
        
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # 写入数据行
        for row_idx, row_data in enumerate(self._rows, 2):
            # 地域
            cell = self.ws.cell(row=row_idx, column=1, value=row_data['region'])
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.thin_border
            
            # 存储类型
            cell = self.ws.cell(row=row_idx, column=2, value=row_data['storage_class'])
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.thin_border
            
            # 冗余类型
            cell = self.ws.cell(row=row_idx, column=3, value=row_data['redundancy_type'])
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.thin_border
            
            # 实际存储（格式化）
            real_storage_str = self._format_storage(row_data['real_storage'])
            cell = self.ws.cell(row=row_idx, column=4, value=real_storage_str)
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.thin_border
            
            # 计费存储（格式化）
            billing_storage_str = self._format_storage(row_data['billing_storage'])
            cell = self.ws.cell(row=row_idx, column=5, value=billing_storage_str)
            cell.font = self.data_font
            cell.alignment = self.data_alignment
            cell.border = self.thin_border
        
        # 生成文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"OSS 使用统计_{timestamp}.xlsx"
        
        # 输出目录（优先读取环境变量，由入口脚本从 config.json 设置）
        output_dir = os.environ.get("ALIYUN_SKILL_OUTPUT_DIR", "/root/.openclaw/workspace/download/")
        os.makedirs(output_dir, exist_ok=True)
        
        file_path = os.path.join(output_dir, filename)
        
        # 保存文件
        self.wb.save(file_path)
        
        return file_path


if __name__ == "__main__":
    # 测试
    generator = OSSExcelGenerator()
    
    # 添加测试数据
    generator.add_row(
        region='cn-hangzhou',
        bucket_count=5,
        storage_class='Standard',
        redundancy_type='LRS',
        real_storage=10737418240,  # 10 GB
        billing_storage=32212254720  # 30 GB (3 副本)
    )
    
    generator.add_row(
        region='cn-hangzhou',
        bucket_count=3,
        storage_class='IA',
        redundancy_type='LRS',
        real_storage=5368709120,  # 5 GB
        billing_storage=16106127360  # 15 GB
    )
    
    generator.add_row(
        region='cn-shanghai',
        bucket_count=2,
        storage_class='Standard',
        redundancy_type='ZRS',
        real_storage=21474836480,  # 20 GB
        billing_storage=64424509440  # 60 GB
    )
    
    # 生成文件
    file_path = generator.generate()
    print(f"测试文件已生成：{file_path}")
